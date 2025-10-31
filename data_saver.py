# data_saver.py

import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.chart import Series
from datetime import datetime
import numpy as np

class DataSaver:
    """
    Classe dedicata al salvataggio dei dati dei test in un file Excel (.xlsx).
    """

    def save_batch_to_xlsx(self, specimens_dict, filepath, calibration_info="N/A"):
        """
        Salva un batch di provini in un singolo file Excel, un provino per foglio.
        """
        try:
            workbook = openpyxl.Workbook()
            # Rimuovi il foglio di default creato automaticamente
            workbook.remove(workbook.active) 

            for specimen_name, specimen_data in specimens_dict.items():
                if specimen_data.get("test_data"): # Salva solo se ci sono dati di test
                    self._create_sheet_for_specimen(workbook, specimen_name, specimen_data, calibration_info)
            
            workbook.save(filepath)
            return True, f"Dati salvati con successo in {filepath}"
        except Exception as e:
            return False, f"Errore durante il salvataggio del file: {e}"

        # In data_saver.py, sostituisci il vecchio _create_sheet_for_specimen con questo:

    def _create_sheet_for_specimen(self, workbook, specimen_name, specimen_data, calibration_info):
        """
        Metodo privato per creare e popolare un singolo foglio di lavoro.
        Ora gestisce sia test Monotonici che Ciclici CON STILE CORRETTO.
        """
        # Controlla se è un test ciclico cercando la chiave speciale che abbiamo aggiunto
        is_cyclic = "test_sequence_setup" in specimen_data

        # --- Creazione Foglio (invariata) ---
        safe_sheet_name = "".join(c for c in specimen_name if c.isalnum() or c in " _-").strip()[:31]
        sheet = workbook.create_sheet(title=safe_sheet_name)

        # --- Scrittura Parametri di Setup (invariata) ---
        sheet["A1"] = "Test Parameters"
        sheet["A1"].font = openpyxl.styles.Font(bold=True)

        params = {
            "Specimen Name": specimen_name,
            "Test Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Calibration Info": calibration_info, 
            "Gauge Length (mm)": specimen_data.get("gauge_length"),
            "Area (mm²)": specimen_data.get("area"),
        }

        row = 2
        for key, value in params.items():
            sheet[f"A{row}"] = key
            sheet[f"B{row}"] = value
            row += 1

        if is_cyclic:
            sheet[f"A{row}"] = "Test Sequence"
            sheet[f"A{row}"].font = openpyxl.styles.Font(bold=True)
            row += 1
            test_sequence = specimen_data.get("test_sequence_setup", [])
            for i, block in enumerate(test_sequence):
                description = self._format_block_description(block, i)
                sheet[f"B{row}"] = description
                row += 1
        else:
            params_monotonic = {
                "Test Speed": f"{specimen_data.get('speed')} {specimen_data.get('speed_unit')}",
                "Stop Criterion": f"{specimen_data.get('stop_criterion_value')} {specimen_data.get('stop_criterion_unit')}"
            }
            for key, value in params_monotonic.items():
                sheet[f"A{row}"] = key
                sheet[f"B{row}"] = value
                row += 1

        # --- Preparazione e Scrittura Colonne Dati (invariata) ---
        header_row = row + 1
        data_start_row = header_row + 1

        headers = [
            "Time (s)", "Relative Displacement (mm)", "Relative Load (N)",
            "Strain (%)", "Stress (MPa)",
            "Absolute Displacement (mm)", "Absolute Load (N)"
        ]

        headers.append("Resistance (Ohm)")

        if is_cyclic:
            headers.extend(["Cycle", "Block"])
        sheet.append(headers)

        test_data = specimen_data.get("test_data", [])
        area = specimen_data.get("area")
        gauge = specimen_data.get("gauge_length")

        for i, data_row in enumerate(test_data):
            resistance = np.nan
            cycle = np.nan
            block = np.nan
            time_s, rel_disp, rel_load, abs_disp, abs_load = data_row[:5]

            if is_cyclic:
                # Tupla ciclica: (... cycle, block, resistance) - 8 elementi
                if len(data_row) == 8: # <-- CONTROLLA LUNGHEZZA
                    cycle = data_row[5]
                    block = data_row[6]
                    resistance = data_row[7] # <-- ESTRAE RESISTENZA DALL'INDICE 7
            else: # Monotonico
                # Tupla monotonica: (... resistance) - 6 elementi
                if len(data_row) == 6: # <-- CONTROLLA LUNGHEZZA
                    resistance = data_row[5] # <-- ESTRAE RESISTENZA DALL'INDICE 5

            is_gauge_valid = isinstance(gauge, (int, float)) and not np.isnan(gauge) and gauge > 0
            is_area_valid = isinstance(area, (int, float)) and not np.isnan(area) and area > 0

            strain = (rel_disp / gauge) * 100 if is_gauge_valid else np.nan
            stress = rel_load / area if is_area_valid else np.nan

            current_excel_row = data_start_row + i
            sheet.cell(row=current_excel_row, column=1, value=time_s)
            sheet.cell(row=current_excel_row, column=2, value=rel_disp)
            sheet.cell(row=current_excel_row, column=3, value=rel_load)
            sheet.cell(row=current_excel_row, column=4, value=strain)
            sheet.cell(row=current_excel_row, column=5, value=stress)
            sheet.cell(row=current_excel_row, column=6, value=abs_disp)
            sheet.cell(row=current_excel_row, column=7, value=abs_load)
            sheet.cell(row=current_excel_row, column=8, value=resistance)
            if is_cyclic:
                sheet.cell(row=current_excel_row, column=9, value=cycle)
                sheet.cell(row=current_excel_row, column=10, value=block)

        num_data_points = len(test_data)
        if num_data_points == 0:
            return 

        # --- Creazione Grafici (CON STILE CORRETTO) ---

        data_end_row = data_start_row + num_data_points - 1 # Riga dati finale

        if is_cyclic:
            # --- Grafici Ciclici (Tempo vs Disp, Tempo vs Load) ---

            # Grafico 1: Tempo vs. Displacement
            chart1 = ScatterChart()
            chart1.title = "Time vs. Displacement"
            self._style_excel_chart(chart1, "Time (s)", "Relative Displacement (mm)", legend=None) # APPLICA STILE

            x_data_ref1 = Reference(sheet, min_col=1, min_row=data_start_row, max_row=data_end_row)
            y_data_ref1 = Reference(sheet, min_col=2, min_row=data_start_row, max_row=data_end_row)
            series1 = Series(y_data_ref1, xvalues=x_data_ref1, title="Disp")
            line_props_data = LineProperties(solidFill="4F81BD", w=38100) # Blu
            series1.graphicalProperties = GraphicalProperties(ln=line_props_data)
            chart1.series.append(series1)
            sheet.add_chart(chart1, "J2") 

            # Grafico 2: Tempo vs. Load
            chart2 = ScatterChart()
            chart2.title = "Time vs. Load"
            self._style_excel_chart(chart2, "Time (s)", "Relative Load (N)", legend=None) # APPLICA STILE

            x_data_ref2 = Reference(sheet, min_col=1, min_row=data_start_row, max_row=data_end_row)
            y_data_ref2 = Reference(sheet, min_col=3, min_row=data_start_row, max_row=data_end_row)
            series2 = Series(y_data_ref2, xvalues=x_data_ref2, title="Load")
            line_props_data2 = LineProperties(solidFill="C0504D", w=38100) # Rosso
            series2.graphicalProperties = GraphicalProperties(ln=line_props_data2)
            chart2.series.append(series2)
            sheet.add_chart(chart2, "J18") 

        else:
            # --- Grafici Monotonici (Logica Originale CON STILE) ---

            # Grafico 1: Load vs Displacement
            chart1 = ScatterChart()
            chart1.title = "Load vs. Displacement"
            self._style_excel_chart(chart1, "Relative Displacement (mm)", "Relative Load (N)", legend=None) # APPLICA STILE

            x_data_ref = Reference(sheet, min_col=2, min_row=data_start_row, max_row=data_end_row)
            y_data_ref = Reference(sheet, min_col=3, min_row=data_start_row, max_row=data_end_row)
            series1 = Series(y_data_ref, xvalues=x_data_ref, title="Test Data")
            line_props1 = LineProperties(solidFill="4F81BD", w=38100) # Blu
            series1.graphicalProperties = GraphicalProperties(ln=line_props1)
            chart1.series.append(series1)
            sheet.add_chart(chart1, "H2") 

            # Grafico 2: Stress vs Strain
            chart2 = ScatterChart()
            chart2.title = "Stress vs. Strain"
            self._style_excel_chart(chart2, "Strain (%)", "Stress (MPa)", legend=None) # APPLICA STILE

            x_data_ref2 = Reference(sheet, min_col=4, min_row=data_start_row, max_row=data_end_row)
            y_data_ref2 = Reference(sheet, min_col=5, min_row=data_start_row, max_row=data_end_row)
            series2 = Series(y_data_ref2, xvalues=x_data_ref2, title="Test Data")
            series2.graphicalProperties = GraphicalProperties(ln=line_props1) # Stesso colore blu
            chart2.series.append(series2)
            sheet.add_chart(chart2, "H18")

    # Inserisci questo metodo dentro la classe DataSaver
    def _format_block_description(self, block, index):

        """ Formatta la descrizione testuale di un blocco. """
        try:
            if block["type"] == "cyclic":
                control_text = block["control_text"]
                if "Displacement" in control_text: unit = "mm"
                elif "Strain" in control_text: unit = "%"
                elif "Force" in control_text: unit = "N"
                else: unit = "MPa"
                description = (
                    f"Block {index+1}: {block['control']} Cycle "
                    f"[{block['lower']:.2f} ↔ {block['upper']:.2f} {unit}] "
                    f"@ {block['speed']:.2f} {block['speed_unit']}, "
                    f"{block['cycles']} cycles "
                    f"(Hold U/L: {block['hold_upper']:.1f}s / {block['hold_lower']:.1f}s)"
                )
            elif block["type"] == "pause":
                description = f"Block {index+1}: Pause [{block['duration']:.1f} s]"
            elif block["type"] == "ramp":
                control_text = block["control_text"]
                if "Displacement" in control_text: unit = "mm"
                elif "Strain" in control_text: unit = "%"
                elif "Force" in control_text: unit = "N"
                else: unit = "MPa"
                hold_str = f", Hold {block['hold_duration']:.1f}s" if block['hold_duration'] > 0 else ""
                description = (
                    f"Block {index+1}: Ramp to {block['target']:.2f} {unit} "
                    f"@ {block['speed']:.2f} {block['speed_unit']}{hold_str}"
                )    
            else:
                description = f"Block {index+1}: Unknown Type"
            return description
        except Exception as e:
            return f"Block {index+1}: Error formatting block data ({e})"
        
        # Inserisci questo metodo dentro la classe DataSaver
# (ad esempio, dopo _format_block_description)

    def _style_excel_chart(self, chart, x_title, y_title, legend=None):
        """ Applica uno stile coerente a un grafico Excel. """
        # chart.title è già impostato prima di chiamare questa funzione
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title
        chart.legend = legend

        # --- Impostazioni di stile dall'originale ---
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.y_axis.majorGridlines = ChartLines() # Griglia solo su Y
        chart.x_axis.majorGridlines = None        # Niente griglia su X
        chart.x_axis.titleOverlay = False
        chart.y_axis.titleOverlay = False

        # --- Stile linea assi (Linea Nera) ---
        line_props_axis = LineProperties(solidFill="000000")
        if chart.x_axis.graphicalProperties is None:
            chart.x_axis.graphicalProperties = GraphicalProperties()    
        chart.x_axis.graphicalProperties.ln = line_props_axis

        if chart.y_axis.graphicalProperties is None:
            chart.y_axis.graphicalProperties = GraphicalProperties()
        chart.y_axis.graphicalProperties.ln = line_props_axis
        # --- Fine impostazioni di stile ---